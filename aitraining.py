# ===== training.py =====
import os
import io
import json
import time
import sys
import math
import numpy as np
import pandas as pd
import lightgbm as lgb
import matplotlib.pyplot as plt
import seaborn as sns
import os
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder
from sklearn.metrics import accuracy_score, precision_recall_fscore_support, classification_report, confusion_matrix
from lime.lime_tabular import LimeTabularExplainer
from matplotlib.ticker import MaxNLocator
from sklearn.utils import resample

import warnings
warnings.filterwarnings("ignore")

OUTPUT_BASE = "/content/drive/MyDrive/maindata/output2"
BUNDLES_ROOT = f"{OUTPUT_BASE}/bundles"

CASES = {
    "UAVCase1": {
        "label_file": "/content/drive/MyDrive/maindata/uavcase1/UAV-Case1-Label.xlsx",
        "merged_csv": "/content/drive/MyDrive/maindata/uavcase1/merged_uavcase1.csv",
    },
    "GCSCase3": {
        "label_file": "/content/drive/MyDrive/maindata/gcscase3/GSCCase3Label.xlsx",
        "merged_csv": "/content/drive/MyDrive/maindata/gcscase3/merged_gcscase3.csv",
        "alt_files": [
            "/content/drive/MyDrive/maindata/gcscase3/GCSCase3Label.xlsx",
            "/content/drive/MyDrive/maindata/gcscase3/GCS-Case3-Label.xlsx",
            "/content/drive/MyDrive/maindata/gcscase3/GCS-Case-3-Label.xlsx",
            "/content/drive/MyDrive/maindata/gcscase3/GCS_Case3_Label.xlsx"
        ]
    },
    "Access": {
        "label_file": "/content/drive/MyDrive/maindata/access/AccessPointCase2Label.xlsx",
        "merged_csv": "/content/drive/MyDrive/maindata/access/merged_access_COMPLETE.csv",
        "alt_files": [
            "/content/drive/MyDrive/maindata/access/merged_access_point.csv",
            "/content/drive/MyDrive/maindata/access/merged_access_PROPER.csv",
            "/content/drive/MyDrive/maindata/access/Access-Label.xlsx",
            "/content/drive/MyDrive/maindata/access/Access-Case-Label.xlsx"
        ]
    }
}

def _print_progress_bar(pct, prefix="", suffix="", width=40):
    pct = 0.0 if pct is None else float(pct)
    pct = max(0.0, min(100.0, pct))
    filled = int(width * pct / 100.0)
    bar = "█" * filled + "░" * (width - filled)
    sys.stdout.write(f"\r{prefix} |{bar}| {pct:6.2f}% {suffix}")
    sys.stdout.flush()
    if pct >= 100.0:
        sys.stdout.write("\n")
        sys.stdout.flush()

def lgb_progress_callback(total_rounds):
    def _callback(env):
        current = getattr(env, "iteration", None)
        if current is None or total_rounds is None or total_rounds <= 0:
            return
        pct = (min(current + 1, total_rounds) / float(total_rounds)) * 100.0
        _print_progress_bar(pct, prefix="Training", suffix=f"({current + 1}/{total_rounds})")
    _callback.order = 10
    return _callback

def _count_lines_fast(path, chunk_bytes=8 * 1024 * 1024):
    total = 0
    with open(path, "rb") as f:
        while True:
            buf = f.read(chunk_bytes)
            if not buf:
                break
            total += buf.count(b"\n")
    return total

def read_csv_with_progress(path, chunksize=200000, dtype=None):
    total_lines = _count_lines_fast(path)
    header_lines = 1 if total_lines > 0 else 0
    denom = max(1, total_lines - header_lines)
    acc = []
    processed = 0
    _print_progress_bar(0, prefix="Loading CSV", suffix="(0 rows)")
    for chunk in pd.read_csv(path, chunksize=chunksize, dtype=dtype):
        acc.append(chunk)
        processed += len(chunk)
        pct = min(100.0, processed / denom * 100.0)
        _print_progress_bar(pct, prefix="Loading CSV", suffix=f"({processed:,} rows)")
    _print_progress_bar(100, prefix="Loading CSV", suffix=f"({processed:,} rows)")
    if acc:
        return pd.concat(acc, axis=0, ignore_index=True)
    return pd.DataFrame()

def read_excel_with_progress(path, sheet_name=0):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[sheet_name] if isinstance(sheet_name, int) else sheet_name]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            _print_progress_bar(100, prefix="Loading Excel", suffix="(0 rows)")
            return pd.DataFrame()
        try:
            total = max(0, ws.max_row - 1)
        except Exception:
            total = None
        data = []
        processed = 0
        last_tick = time.time()
        _print_progress_bar(0, prefix="Loading Excel", suffix="(0 rows)")
        for r in rows:
            data.append(r)
            processed += 1
            if time.time() - last_tick > 0.1:
                if total and total > 0:
                    pct = min(100.0, processed / total * 100.0)
                    _print_progress_bar(pct, prefix="Loading Excel", suffix=f"({processed:,} rows)")
                else:
                    _print_progress_bar(None, prefix="Loading Excel", suffix=f"({processed:,} rows)")
                last_tick = time.time()
        _print_progress_bar(100, prefix="Loading Excel", suffix=f"({processed:,} rows)")
        df = pd.DataFrame(data, columns=list(header))
        wb.close()
        return df
    except Exception:
        _print_progress_bar(0, prefix="Loading Excel", suffix="(start)")
        df = pd.read_excel(path, sheet_name=sheet_name)
        _print_progress_bar(100, prefix="Loading Excel", suffix=f"({len(df):,} rows)")
        return df

def setup_complete_output_structure():
    folders = [
        OUTPUT_BASE,
        f"{OUTPUT_BASE}/plots",
        f"{OUTPUT_BASE}/confusionmatrices",
        f"{OUTPUT_BASE}/reports",
        f"{OUTPUT_BASE}/lime_reports",
        f"{OUTPUT_BASE}/lime_reports/uavcase1",
        f"{OUTPUT_BASE}/lime_reports/gcscase3",
        f"{OUTPUT_BASE}/lime_reports/access",
        BUNDLES_ROOT
    ]
    for folder in folders:
        os.makedirs(folder, exist_ok=True)
    return folders

def check_available_files():
    base_dirs = {
        "UAVCase1": "/content/drive/MyDrive/maindata/uavcase1",
        "GCSCase3": "/content/drive/MyDrive/maindata/gcscase3",
        "Access": "/content/drive/MyDrive/maindata/access"
    }
    for case_name, directory in base_dirs.items():
        print(f"{case_name} ({directory})")
        if os.path.exists(directory):
            try:
                files = os.listdir(directory)
                label_files = [f for f in files if 'label' in f.lower() and (f.endswith('.xlsx') or f.endswith('.csv'))]
                merged_files = [f for f in files if 'merged' in f.lower() and (f.endswith('.csv') or f.endswith('.xlsx'))]
                other_data_files = [f for f in files if f.endswith(('.csv', '.xlsx')) and f not in label_files + merged_files]
                print(f"  label: {label_files if label_files else 'none'}")
                print(f"  merged: {merged_files if merged_files else 'none'}")
                if other_data_files:
                    print(f"  others: {other_data_files[:3]}")
            except Exception as e:
                print(f"  error: {e}")
        else:
            print("  missing")

def load_case_data_comprehensive(case_name):
    case_config = CASES[case_name]
    files_to_try = []
    if "label_file" in case_config:
        files_to_try.append(("Label File", case_config["label_file"]))
    if "merged_csv" in case_config:
        files_to_try.append(("Merged CSV", case_config["merged_csv"]))
    if "alt_files" in case_config:
        for alt_file in case_config["alt_files"]:
            files_to_try.append(("Alternative", alt_file))
    loaded_data = []
    for file_type, file_path in files_to_try:
        if os.path.exists(file_path):
            try:
                if file_path.endswith('.xlsx'):
                    data = read_excel_with_progress(file_path)
                else:
                    data = read_csv_with_progress(file_path)
                target_col = find_target_column_comprehensive(data, case_name)
                if target_col:
                    target_classes = data[target_col].value_counts()
                    loaded_data.append({
                        'data': data,
                        'source': file_type,
                        'file': os.path.basename(file_path),
                        'target_col': target_col,
                        'num_classes': len(target_classes)
                    })
            except Exception as e:
                pass
    if not loaded_data:
        return None, None
    best_data = max(loaded_data, key=lambda x: (x['num_classes'], len(x['data'])))
    return best_data['data'], best_data['target_col']

def find_target_column_comprehensive(data, case_name):
    primary_targets = ['attack_type', 'Attack_Type', 'Label', 'label', 'class', 'Class']
    for col in primary_targets:
        if col in data.columns:
            classes = data[col].nunique()
            if 2 <= classes <= 50:
                return col
    attack_keywords = ['dos', 'ddos', 'mitm', 'bruteforce', 'brute', 'normal', 'attack', 'scan', 'flood', 'jam']
    for col in data.columns:
        if data[col].dtype == 'object':
            unique_vals = data[col].unique()
            num_classes = len(unique_vals)
            if 2 <= num_classes <= 50:
                val_text = ' '.join(str(v).lower() for v in unique_vals[:20])
                if any(keyword in val_text for keyword in attack_keywords):
                    return col
    last_col = data.columns[-1]
    if data[last_col].dtype == 'object':
        classes = data[last_col].nunique()
        if 2 <= classes <= 50:
            return last_col
    for col in data.columns:
        if data[col].dtype == 'object':
            classes = data[col].nunique()
            if 2 <= classes <= 50:
                return col
    return None

def clean_labels_series(s):
    mapping = {
        "fake landing ": "Fake Landing",
        "fake landing": "Fake Landing",
        "bruteforce": "BruteForce",
        "de-authentication": "Deauthentication",
        "deauthentication": "Deauthentication",
        "evil twin": "Evil Twin",
        "eviltwin": "Evil Twin",
        "reconnassiance": "Reconnaissance",
        "icmp flooding": "ICMP Flooding",
        "udp flooding": "UDP Flooding",
        "gps jamming": "GPS Jamming",
        "benign": "Normal",
        "ddos": "DDoS",
        "dos": "DoS",
    }
    s2 = s.astype(str).str.strip()
    s2 = s2.str.replace(r"\s+", " ", regex=True).str.lower()
    s2 = s2.replace(mapping)
    s2 = s2.str.replace("_", " ").str.title()
    s2 = s2.replace({
        "Ddos": "DDoS",
        "Dos": "DoS",
        "Mitm": "MITM",
        "Udp Flooding": "UDP Flooding",
        "Icmp Flooding": "ICMP Flooding",
        "Gps Jamming": "GPS Jamming",
        "Brute Force": "BruteForce",
    })
    return s2

def get_default_thresholds(class_names):
    high = {"DDoS", "DoS", "MITM", "Jamming", "Fake Landing", "GPS Jamming", "Evil Twin", "UDP Flooding"}
    medium = {"BruteForce", "Deauthentication", "Reconnaissance", "Replay", "ICMP Flooding"}
    thresholds = {}
    for c in class_names:
        if c in {"Normal", "Benign"}:
            thresholds[c] = 9.99
        elif c in high:
            thresholds[c] = 0.70
        elif c in medium:
            thresholds[c] = 0.60
        else:
            thresholds[c] = 0.50
    return thresholds

def prepare_features_maximum_retention(data, target_col, case_name):
    y = clean_labels_series(data[target_col].copy())
    X = data.drop(columns=[target_col]).copy()
    obvious_meta = ['id', 'source_file', 'source_folder', 'filename', 'uid']
    for col in list(X.columns):
        if col.lower() in [m.lower() for m in obvious_meta]:
            X = X.drop(columns=[col])
    numeric_cols = []
    for col in X.columns:
        if X[col].dtype in ['int64', 'float64', 'int32', 'float32']:
            numeric_cols.append(col)
        else:
            X[col] = pd.to_numeric(X[col], errors='coerce')
            if not X[col].isnull().all():
                numeric_cols.append(col)
    X_numeric = X[numeric_cols].copy()
    for col in list(X_numeric.columns):
        missing_count = X_numeric[col].isnull().sum()
        missing_pct = missing_count / len(X_numeric) * 100
        if missing_count > 0:
            if missing_pct > 95:
                X_numeric = X_numeric.drop(columns=[col])
            else:
                median_val = X_numeric[col].median()
                X_numeric[col] = X_numeric[col].fillna(median_val)
    zero_var_cols = [c for c in X_numeric.columns if X_numeric[c].std() == 0]
    if zero_var_cols:
        X_numeric = X_numeric.drop(columns=zero_var_cols)
    for col in X_numeric.columns:
        if np.isinf(X_numeric[col]).any():
            X_numeric[col] = X_numeric[col].replace([np.inf, -np.inf], X_numeric[col].median())
    return X_numeric.astype(np.float32), y

def summarize_counts(y, title="Counts"):
    vc = pd.Series(y).value_counts().sort_values(ascending=False)
    print(f"{title}: " + ", ".join([f"{k}:{v}" for k,v in vc.items()]))

def rebalance_train_set(X_train, y_train, *, max_per_class=20000, min_per_class=2000, upsample_small=True, random_state=42):
    rng = np.random.RandomState(random_state)
    X_list, y_list = [], []
    df = X_train.copy()
    df['__y__'] = y_train
    for cls, group in df.groupby('__y__'):
        n = len(group)
        if n > max_per_class:
            grp = group.sample(n=max_per_class, random_state=rng)
        elif upsample_small and n < min_per_class:
            need = min_per_class - n
            boost = resample(group, replace=True, n_samples=need, random_state=rng)
            grp = pd.concat([group, boost], axis=0)
        else:
            grp = group
        X_list.append(grp.drop(columns=['__y__']))
        y_list.append(np.full(len(grp), cls, dtype=y_train.dtype))
    Xb = pd.concat(X_list, axis=0).reset_index(drop=True)
    yb = np.concatenate(y_list, axis=0)
    idx = np.arange(len(Xb))
    rng.shuffle(idx)
    Xb = Xb.iloc[idx].reset_index(drop=True)
    yb = yb[idx]
    return Xb, yb

def save_fig(path, fig=None):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    if fig is None:
        plt.savefig(path, dpi=300, bbox_inches='tight')
        plt.close()
    else:
        fig.savefig(path, dpi=300, bbox_inches='tight')
        plt.close(fig)

def plot_probabilities_bar(all_classes, probs_vec, case_name, class_name, out_path):
    order = np.argsort(probs_vec)[::-1]
    classes_sorted = [all_classes[i] for i in order]
    probs_sorted = probs_vec[order]
    plt.figure(figsize=(10, max(4, 0.4*len(all_classes))))
    ax = plt.gca()
    ax.barh(range(len(classes_sorted)), probs_sorted, alpha=0.85)
    ax.set_yticks(range(len(classes_sorted)))
    ax.set_yticklabels(classes_sorted)
    ax.invert_yaxis()
    ax.set_xlabel("Probability")
    ax.set_title(f"{case_name} – All class probabilities (instance for {class_name})")
    ax.xaxis.set_major_locator(MaxNLocator(6))
    save_fig(out_path)

def plot_threshold_bar(prob, threshold, case_name, class_name, out_path):
    pct = float(prob)
    thr = float(threshold)
    plt.figure(figsize=(8, 2.2))
    ax = plt.gca()
    ax.barh([0], [pct], height=0.5)
    ax.axvline(thr, color='red', linestyle='--', linewidth=2, label=f"Threshold {thr:.0%}")
    ax.set_xlim(0, 1)
    ax.set_yticks([])
    ax.set_xlabel("Probability")
    ax.set_title(f"{case_name} – {class_name}: probability vs threshold")
    ax.legend(loc='lower right')
    save_fig(out_path)

def plot_lime_bars(exp_pairs, class_name, confidence, case_name, out_path, top_k=12):
    pairs = exp_pairs[:top_k]
    if not pairs:
        plt.figure(figsize=(8,2))
        plt.text(0.5, 0.5, "No LIME features available", ha='center', va='center')
        plt.axis('off')
        save_fig(out_path)
        return
    features = [p[0] for p in pairs]
    values = [p[1] for p in pairs]
    colors = ['green' if v > 0 else 'red' for v in values]
    fig_h = max(6, 0.45 * len(features))
    plt.figure(figsize=(11, fig_h))
    ax = plt.gca()
    ax.barh(range(len(features)), values, color=colors, alpha=0.85)
    short = [f[:70] + '…' if len(f) > 70 else f for f in features]
    ax.set_yticks(range(len(features)))
    ax.set_yticklabels(short)
    ax.invert_yaxis()
    ax.set_xlabel("Contribution to class score")
    ax.set_title(f"{case_name} – LIME contributions for {class_name} (conf {confidence:.1%})")
    ax.grid(True, axis='x', alpha=0.25)
    save_fig(out_path)

def create_enhanced_confusion_matrix(y_true, y_pred, classes, case_name, normalize=False):
    cm = confusion_matrix(y_true, y_pred)
    suffix = "_normalized" if normalize else "_counts"
    filename = f"confusion_matrix_{case_name.lower()}{suffix}.png"
    file_path = os.path.join(OUTPUT_BASE, "confusionmatrices", filename)
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    fig_size = max(8, len(classes) * 0.8), max(6, len(classes) * 0.6)
    plt.figure(figsize=fig_size)
    if normalize:
        cm_display = cm.astype('float') / (cm.sum(axis=1)[:, np.newaxis] + 1e-10)
        fmt = '.2f'; title = f'{case_name} Confusion Matrix (Normalized)'; cbar_label = 'Proportion'
    else:
        cm_display = cm; fmt = 'd'; title = f'{case_name} Confusion Matrix (Counts)'; cbar_label = 'Count'
    sns.heatmap(cm_display, annot=True, fmt=fmt, cmap='Blues',
                xticklabels=classes, yticklabels=classes,
                cbar_kws={'label': cbar_label})
    plt.title(title, fontsize=16, fontweight='bold', pad=20)
    plt.xlabel('Predicted Label', fontsize=14)
    plt.ylabel('True Label', fontsize=14)
    plt.xticks(rotation=45, ha='right')
    plt.yticks(rotation=0)
    plt.tight_layout()
    save_fig(file_path)
    try:
        csv_path = file_path.replace('.png', '.csv')
        pd.DataFrame(cm_display, index=classes, columns=classes).to_csv(csv_path)
    except Exception:
        pass

def create_training_curves(model, case_name):
    if not hasattr(model, 'evals_result_') or not model.evals_result_:
        return None
    res = model.evals_result_
    valid_key = list(res.keys())[0]
    metric_key = list(res[valid_key].keys())[0]
    scores = res[valid_key][metric_key]
    filename = f"training_curves_{case_name.lower()}.png"
    file_path = os.path.join(OUTPUT_BASE, "plots", filename)
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    plt.figure(figsize=(12, 6))
    plt.plot(scores, linewidth=2, label=f'Validation {metric_key}', color='blue')
    best_iter = np.argmin(scores) + 1
    plt.axvline(best_iter, linestyle='--', alpha=0.7, color='red', label=f'Best iteration: {best_iter}')
    plt.xlabel('Iteration', fontsize=12)
    plt.ylabel(metric_key.replace('_', ' ').title(), fontsize=12)
    plt.title(f'{case_name} Training Progress', fontsize=16, fontweight='bold')
    plt.legend()
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    save_fig(file_path)
    try:
        csv_path = file_path.replace('.png', '.csv')
        pd.DataFrame({'iteration': range(len(scores)), 'validation_score': scores}).to_csv(csv_path, index=False)
    except Exception:
        pass
    return file_path

def create_feature_importance_plot(model, feature_names, case_name, top_n=20):
    importance = model.feature_importances_
    feature_df = pd.DataFrame({'feature': feature_names, 'importance': importance}) \
                    .sort_values('importance', ascending=False).head(top_n)
    filename = f"feature_importance_{case_name.lower()}.png"
    file_path = os.path.join(OUTPUT_BASE, "plots", filename)
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    plt.figure(figsize=(12, max(8, len(feature_df) * 0.45)))
    y_pos = np.arange(len(feature_df))
    bars = plt.barh(y_pos, feature_df['importance'], alpha=0.85, color='skyblue')
    plt.yticks(y_pos, feature_df['feature'])
    plt.xlabel('Importance Score', fontsize=12)
    plt.title(f'{case_name} Top {top_n} Feature Importance', fontsize=16, fontweight='bold')
    plt.gca().invert_yaxis()
    for i, (bar, val) in enumerate(zip(bars, feature_df['importance'])):
        plt.text(val + max(feature_df['importance']) * 0.01, i, f'{val:.0f}', va='center', fontsize=10)
    plt.tight_layout()
    save_fig(file_path)
    try:
        csv_path = file_path.replace('.png', '.csv')
        feature_df.to_csv(csv_path, index=False)
    except Exception:
        pass
    return file_path

def create_lime_pngs_only(X_train, X_test, y_train, y_test, model, label_encoder, case_name):
    lime_folder = os.path.join(OUTPUT_BASE, "lime_reports", case_name.lower())
    os.makedirs(lime_folder, exist_ok=True)
    explainer = LimeTabularExplainer(
        X_train.values,
        feature_names=X_train.columns.tolist(),
        class_names=label_encoder.classes_.tolist(),
        discretize_continuous=True,
        mode='classification',
        random_state=42
    )
    class_names = label_encoder.classes_.tolist()
    thresholds = get_default_thresholds(class_names)
    outputs = []
    all_probs = model.predict_proba(X_test)
    max_confidences = np.max(all_probs, axis=1)
    total_classes = len(class_names)
    for class_idx, class_name in enumerate(class_names):
        pct = ((class_idx) / max(1, total_classes)) * 100.0
        _print_progress_bar(pct, prefix="LIME per-class", suffix=f"({class_idx}/{total_classes})")
        class_indices = np.where(y_test == class_idx)[0]
        if len(class_indices) == 0:
            continue
        ci = class_indices
        class_confidences = max_confidences[ci]
        best_idx = ci[np.argmax(class_confidences)]
        probs_vec = all_probs[best_idx]
        this_prob = float(probs_vec[class_idx])
        pred_idx = int(np.argmax(probs_vec))
        pred_name = class_names[pred_idx]
        confidence = float(np.max(probs_vec))
        explanation = explainer.explain_instance(
            X_test.iloc[best_idx].values,
            model.predict_proba,
            num_features=15,
            labels=[class_idx]
        )
        exp_pairs = explanation.as_list(label=class_idx)
        lime_png = os.path.join(lime_folder, f"lime_{class_name.replace(' ', '_')}.png")
        plot_lime_bars(exp_pairs, class_name, confidence, case_name, lime_png, top_k=12)
        probs_png = os.path.join(lime_folder, f"probs_{class_name.replace(' ', '_')}.png")
        plot_probabilities_bar(class_names, probs_vec, case_name, class_name, probs_png)
        thr_png = None
        if class_name not in {"Normal", "Benign"}:
            thr = thresholds.get(class_name, 0.5)
            thr_png = os.path.join(lime_folder, f"threshold_{class_name.replace(' ', '_')}.png")
            plot_threshold_bar(this_prob, thr, case_name, class_name, thr_png)
        outputs.append({
            "class": class_name,
            "instance_index": int(best_idx),
            "predicted_class": pred_name,
            "prob_for_class": this_prob,
            "confidence": confidence,
            "lime_png": lime_png,
            "probs_png": probs_png,
            "threshold_png": thr_png if thr_png else "",
            "threshold": thresholds.get(class_name, 0.5)
        })
    _print_progress_bar(100, prefix="LIME per-class", suffix=f"({total_classes}/{total_classes})")
    if outputs:
        df = pd.DataFrame(outputs)
        csv_path = os.path.join(lime_folder, f"lime_png_summary_{case_name.lower()}.csv")
        df.to_csv(csv_path, index=False)
    return outputs

def export_bundle(case_name, model, label_encoder, feature_names, train_feature_medians, thresholds):
    out_dir = os.path.join(BUNDLES_ROOT, case_name.lower())
    os.makedirs(out_dir, exist_ok=True)
    import joblib
    joblib.dump(model, os.path.join(out_dir, "model.pkl"))
    joblib.dump(label_encoder, os.path.join(out_dir, "label_encoder.pkl"))
    spec = {
        "class_names": label_encoder.classes_.tolist(),
        "feature_names": list(feature_names),
        "feature_medians": {k: float(train_feature_medians[k]) for k in feature_names},
        "thresholds": {k: float(thresholds.get(k, 0.5)) for k in label_encoder.classes_.tolist()}
    }
    with open(os.path.join(out_dir, "feature_spec.json"), "w") as f:
        json.dump(spec, f, indent=2)
    return out_dir
def _cr_to_dataframe(y_true, y_pred, class_names):
    rep = classification_report(y_true, y_pred, target_names=class_names, output_dict=True, zero_division=0)
    rows = []
    for cls in class_names:
        if cls in rep:
            r = rep[cls]
            rows.append({
                "Class": cls,
                "Precision": float(r.get("precision", 0.0)),
                "Recall": float(r.get("recall", 0.0)),
                "F1-Score": float(r.get("f1-score", 0.0)),
                "Support": int(r.get("support", 0))
            })

    df = pd.DataFrame(rows)
    df["__order"] = df["Class"].apply(lambda c: class_names.index(c))
    df = df.sort_values("__order").drop(columns="__order")
    return df
def create_training_curves_dual(model, case_name, out_dir):
    res = getattr(model, "evals_result_", None)
    if not res: return None, None, None
    ds = list(res.keys())
    if len(ds) < 2: return None, None, None
    metric = list(res[ds[0]].keys())[0]
    train_curve = res[ds[0]][metric]
    valid_curve = res[ds[1]][metric]
    best_iter = int(np.argmin(valid_curve)) + 1
    os.makedirs(out_dir, exist_ok=True)
    p1 = os.path.join(out_dir, f"training_curves_dual_{case_name.lower()}.png")
    plt.figure(figsize=(12,6))
    plt.plot(train_curve, label="Train")
    plt.plot(valid_curve, label="Validation")
    plt.axvline(best_iter, linestyle="--")
    plt.xlabel("Iteration"); plt.ylabel(metric); plt.title(f"{case_name} Training vs Validation")
    plt.legend(); plt.grid(True, alpha=.3); plt.tight_layout(); plt.savefig(p1, dpi=300); plt.close()
    return p1, train_curve, valid_curve

def create_training_curves_zoom(case_name, train_curve, valid_curve, out_dir, window=60):
    if train_curve is None or valid_curve is None: return None
    best_iter = int(np.argmin(valid_curve))
    half = max(window//2, 20)
    lo = max(0, best_iter - half)
    hi = min(len(valid_curve), best_iter + half)
    x = np.arange(lo, hi)
    os.makedirs(out_dir, exist_ok=True)
    p2 = os.path.join(out_dir, f"training_curves_zoom_{case_name.lower()}.png")
    plt.figure(figsize=(12,6))
    plt.plot(x, np.array(train_curve)[lo:hi], label="Train")
    plt.plot(x, np.array(valid_curve)[lo:hi], label="Validation")
    plt.axvline(best_iter, linestyle="--")
    plt.xlabel("Iteration"); plt.ylabel("score"); plt.title(f"{case_name} Zoomed Validation Neighborhood")
    plt.legend(); plt.grid(True, alpha=.3); plt.tight_layout(); plt.savefig(p2, dpi=300); plt.close()
    return p2

def _save_classification_report_csv_png(y_true, y_pred, class_names, case_name, base_dir):
    os.makedirs(base_dir, exist_ok=True)
    df = _cr_to_dataframe(y_true, y_pred, class_names)
    csv_path = os.path.join(base_dir, f"classification_report_{case_name.lower()}.csv")
    df.to_csv(csv_path, index=False)

   
    plt.figure(figsize=(max(8, 0.4 * len(df) * 4), 0.6 * (len(df) + 4)))
    ax = plt.gca()
    tbl = ax.table(
        cellText=np.round(df[["Precision","Recall","F1-Score"]].values, 3),
        rowLabels=df["Class"].tolist(),
        colLabels=["Precision", "Recall", "F1-Score"],
        cellLoc="center",
        rowLoc="center",
        loc="center"
    )
    ax.axis("off")
    ax.set_title(f"{case_name} — Per-Class Detection Performance", pad=18, fontsize=14)
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(9)
    tbl.scale(1, 1.2)
    png_path = os.path.join(base_dir, f"classification_report_{case_name.lower()}.png")
    plt.tight_layout()
    plt.savefig(png_path, dpi=300, bbox_inches="tight")
    plt.close()
    print(f"  Saved: {os.path.basename(csv_path)}")
    print(f"  Saved: {os.path.basename(png_path)}")

def train_comprehensive_model(X, y, case_name):
    setup_complete_output_structure()
    class_counts = y.value_counts()
    rare_threshold = 2
    rare_classes = class_counts[class_counts < rare_threshold].index.tolist()
    if rare_classes:
        mask = ~y.isin(rare_classes)
        X_filtered = X[mask].copy()
        y_filtered = y[mask].copy()
    else:
        X_filtered = X.copy()
        y_filtered = y.copy()
    le = LabelEncoder()
    y_encoded = le.fit_transform(y_filtered)
    X_train_full, X_test, y_train_full, y_test = train_test_split(
        X_filtered, y_encoded, test_size=0.2, random_state=42,
        stratify=y_encoded if len(np.unique(y_encoded)) > 1 else None
    )
    X_train, X_val, y_train, y_val = train_test_split(
        X_train_full, y_train_full, test_size=0.2, random_state=42,
        stratify=y_train_full if len(np.unique(y_train_full)) > 1 else None
    )
    summarize_counts(y_train, "Before train balance")
    X_train, y_train = rebalance_train_set(
        X_train, y_train,
        max_per_class=20000,
        min_per_class=2000,
        upsample_small=True,
        random_state=42
    )
    summarize_counts(y_train, "After  train balance")
    lgb_params = {
        'objective': 'multiclass' if len(le.classes_) > 2 else 'binary',
        'num_class': len(le.classes_) if len(le.classes_) > 2 else None,
        'learning_rate': 0.05,
        'n_estimators': 4000,
        'num_leaves': 63,
        'max_depth': -1,
        'min_child_samples': 40,
        'subsample': 0.8,
        'subsample_freq': 1,
        'colsample_bytree': 0.8,
        'reg_alpha': 0.2,
        'reg_lambda': 0.2,
        'max_bin': 255,
        'random_state': 42,
        'n_jobs': -1,
        'class_weight': None
    }
    model = lgb.LGBMClassifier(**lgb_params)
    start_time = time.time()
    model.fit(
        X_train, y_train,
        eval_set=[(X_train, y_train), (X_test, y_test)],
        eval_metric='multi_logloss' if len(le.classes_) > 2 else 'binary_logloss',
        callbacks=[lgb_progress_callback(lgb_params.get('n_estimators', 0))]
    )
    plots_dir = os.path.join(OUTPUT_BASE, "plots")
    p_all, tr, va = create_training_curves_dual(model, case_name, plots_dir)
    p_zoom = create_training_curves_zoom(case_name, tr, va, plots_dir, window=80)

    training_time = time.time() - start_time
    y_pred = model.predict(X_test, num_iteration=model.best_iteration_)
    y_pred_proba = model.predict_proba(X_test, num_iteration=model.best_iteration_)
    accuracy = accuracy_score(y_test, y_pred)
    precision, recall, f1, _ = precision_recall_fscore_support(
        y_test, y_pred, average='weighted', zero_division=0
    )
    print("\nDETAILED CLASSIFICATION REPORT:")
    print(classification_report(y_test, y_pred, target_names=le.classes_, digits=4, zero_division=0))
    
    report_dir = os.path.join(OUTPUT_BASE, "reports")
    _save_classification_report_csv_png(
        y_true=y_test,
        y_pred=y_pred,
        class_names=le.classes_.tolist(),
        case_name=case_name,
        base_dir=report_dir
    )
    best_iter = int(getattr(model, "best_iteration_", model.get_params().get('n_estimators', 0)))
    print(f"best_iteration: {best_iter}")
    print(f"accuracy: {accuracy:.4f}")
    print(f"f1_score: {f1:.4f}")
    create_enhanced_confusion_matrix(y_test, y_pred, le.classes_, case_name, normalize=False)
    create_enhanced_confusion_matrix(y_test, y_pred, le.classes_, case_name, normalize=True)
    create_training_curves(model, case_name)
    create_feature_importance_plot(model, X_filtered.columns.tolist(), case_name)
    lime_pngs = create_lime_pngs_only(
        pd.DataFrame(X_train, columns=X_filtered.columns),
        pd.DataFrame(X_test, columns=X_filtered.columns),
        y_train, y_test, model, le, case_name
    )
    results_summary = {
        'case_name': case_name,
        'timestamp': time.strftime('%Y-%m-%d %H:%M:%S'),
        'data_info': {
            'total_samples': len(X_filtered),
            'features': X_filtered.shape[1],
            'classes': len(le.classes_),
            'class_names': le.classes_.tolist(),
            'removed_rare_classes': rare_classes
        },
        'performance': {
            'accuracy': float(accuracy),
            'precision': float(precision),
            'recall': float(recall),
            'f1_score': float(f1),
            'training_time': float(training_time),
            'best_iteration': best_iter
        }
    }
    summary_path = os.path.join(OUTPUT_BASE, "reports", f"{case_name.lower()}_complete_results.json")
    os.makedirs(os.path.dirname(summary_path), exist_ok=True)
    with open(summary_path, 'w') as f:
        json.dump(results_summary, f, indent=2, default=str)
    train_medians = X_train.median().astype(float)
    thresholds = get_default_thresholds(le.classes_.tolist())
    bundle_dir = export_bundle(case_name, model, le, X_filtered.columns.tolist(), train_medians, thresholds)
    print(f"bundle: {bundle_dir}")
    print(f"report: {summary_path}")
    return {
        'model': model,
        'label_encoder': le,
        'feature_names': X_filtered.columns.tolist(),
        'classes': le.classes_.tolist(),
        'accuracy': accuracy,
        'f1_score': f1,
        'training_time': training_time,
        'best_iteration': best_iter,
        'lime_pngs': lime_pngs,
        'results_summary': results_summary
    }

def train_uavcase1():
    data, target_col = load_case_data_comprehensive("UAVCase1")
    if data is None:
        return None
    X, y = prepare_features_maximum_retention(data, target_col, "UAVCase1")
    return train_comprehensive_model(X, y, "UAVCase1")

def train_gcscase3():
    data, target_col = load_case_data_comprehensive("GCSCase3")
    if data is None:
        return None
    X, y = prepare_features_maximum_retention(data, target_col, "GCSCase3")
    return train_comprehensive_model(X, y, "GCSCase3")

def train_access():
    data, target_col = load_case_data_comprehensive("Access")
    if data is None:
        return None
    X, y = prepare_features_maximum_retention(data, target_col, "Access")
    return train_comprehensive_model(X, y, "Access")

def train_all_cases():
    setup_complete_output_structure()
    all_results = {}
    cases = {
        'UAVCase1': train_uavcase1,
        'GCSCase3': train_gcscase3,
        'Access': train_access
    }
    for case_name, train_func in cases.items():
        try:
            results = train_func()
            all_results[case_name] = results
        except Exception as e:
            all_results[case_name] = None
    return all_results

print("check_available_files()")
print("uav_results = train_uavcase1()")
print("gcs_results = train_gcscase3()")
print("access_results = train_access()")
print("all_results = train_all_cases()")
print(f"LIME PNGs: {OUTPUT_BASE}/lime_reports/<case>/")
